#!/usr/bin/env python3
"""Export activity-log.md -> activity-log.xlsx, formatted for handing in.

Header band, per-status row colours, numeric Hours column with a TOTAL row,
frozen header and an autofilter. Requires openpyxl (`pip install openpyxl`);
if you don't have it, use export_csv.py instead.

Usage:
  python export_excel.py                       # paths from timesheet-config.json
  python export_excel.py --src X.md --out Y.xlsx
"""
import argparse
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from tsconfig import data_root, log_path  # noqa: E402

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl is not installed.  pip install openpyxl   (or use export_csv.py)")
    raise SystemExit(1)

HEADERS = ["Date", "Project", "Task", "Hours", "Status", "Notes"]


def parse_rows(src):
    rows = []
    with open(src, encoding="utf-8") as fh:
        for line in fh:
            s = line.strip()
            if not s.startswith("|"):
                continue
            cells = [c.strip() for c in s.strip("|").split("|")]
            if len(cells) != 6:
                continue
            if cells[0] == "Date" or set(cells[0]) <= set("-: "):
                continue  # header or separator
            rows.append(cells)
    rows.sort(key=lambda r: r[0])  # by date
    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", default=log_path())
    ap.add_argument("--out", default=os.path.join(data_root(), "activity-log.xlsx"))
    a = ap.parse_args()

    if not os.path.exists(a.src):
        print(f"no log at {a.src} - is dataRoot set correctly? (python tsconfig.py)")
        return

    rows = parse_rows(a.src)

    wb = Workbook()
    ws = wb.active
    ws.title = "Activity Log"

    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="305496")
    hdr_font = Font(bold=True, color="FFFFFF")
    status_fill = {
        "Done": PatternFill("solid", fgColor="E2EFDA"),
        "In progress": PatternFill("solid", fgColor="FFF2CC"),
        "Blocked": PatternFill("solid", fgColor="FCE4D6"),
    }

    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, c, h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    total = 0.0
    for r, data in enumerate(rows, 2):
        for c, val in enumerate(data, 1):
            cell = ws.cell(r, c, val)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (3, 6)))
        try:  # hours numeric
            h = float(data[3])
            ws.cell(r, 4).value = h
            total += h
        except ValueError:
            pass
        f = status_fill.get(data[4])
        if f:
            ws.cell(r, 5).fill = f

    tr = len(rows) + 2
    ws.cell(tr, 3, "TOTAL").font = Font(bold=True)
    ws.cell(tr, 3).alignment = Alignment(horizontal="right")
    tc = ws.cell(tr, 4, round(total, 2))
    tc.font = Font(bold=True)
    tc.fill = PatternFill("solid", fgColor="DDEBF7")

    for c, w in {1: 12, 2: 13, 3: 70, 4: 8, 5: 13, 6: 40}.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{len(rows) + 1}"

    wb.save(a.out)
    print(f"Wrote {a.out}")
    print(f"{len(rows)} task rows, total hours = {round(total, 2)}")


if __name__ == "__main__":
    main()
